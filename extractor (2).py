# ── extractor.py ─────────────────────────────────────────────────────────
# Extrae datos de PDFs y Excels usando la API de Anthropic
# ─────────────────────────────────────────────────────────────────────────
import base64, json, re
import anthropic
import openpyxl
from collections import defaultdict

SYSTEM_PROMPT = """Sos un asistente especializado en extraer datos financieros de estados de cuenta.
Respondé ÚNICAMENTE con JSON válido, sin texto adicional, sin markdown, sin backticks.

Para statements ANUALES de Pershing (Year-End Account Report):
{
  "type": "annual",
  "custodian": "pershing",
  "years": {
    "2024": {"vi": 0, "dep": 0, "ret": 0, "vf": 0, "res": 0, "r": 0.0},
    "2025": {"vi": 0, "dep": 0, "ret": 0, "vf": 0, "res": 0, "r": 0.0}
  }
}

Para statements MENSUALES de Pershing (Monthly Statement):
{
  "type": "monthly",
  "custodian": "pershing",
  "year": 2026,
  "month": 4,
  "label": "YTD 2026",
  "vi": 0, "dep": 0, "ret": 0, "vf": 0, "res": 0, "r": 0.0
}

Para statements de IBKR (Activity Statement):
{
  "type": "monthly",
  "custodian": "ibkr",
  "year": 2026,
  "month": 5,
  "label": "YTD 2026",
  "vi": 0, "dep": 0, "ret": 0, "vf": 0, "res": 0, "r": 0.0,
  "positions": [
    {"symbol": "", "description": "", "secType": "ETF", "assetCategory": "Alternativo",
     "geo": "Global", "sector": "", "quantity": 0, "price": 0.0,
     "marketValue": 0.0, "costBasis": 0.0}
  ]
}

Para statements de STONEX:
{
  "type": "monthly",
  "custodian": "stonex",
  "year": 2026,
  "month": 4,
  "label": "YTD 2026",
  "vi": 0, "dep": 0, "ret": 0, "vf": 0, "res": 0, "r": 0.0,
  "positions": []
}

NOTAS IMPORTANTES:
- "vi" = valor inicio del período
- "dep" = depósitos en efectivo y títulos (positivo)
- "ret" = retiros en efectivo y títulos + debit card activity (positivo)
- "vf" = valor final del período
- "res" = resultado = vf - vi - dep + ret
- "r" = retorno MWR = vf / (vi + dep*0.5 - ret*0.5) - 1
- Para Pershing mensual YTD: vi = Beginning Account Value (Jan 1), vf = Ending Account Value
- Para IBKR: vi = NAV inicio, vf = NAV final, ret incluye comisiones del asesor
- Para retornos anuales históricos en informe anual de Pershing: extraer todos los años disponibles
- assetCategory debe ser uno de: Renta Variable, Renta Fija, Alternativo, Fondos, Efectivo
- geo debe ser uno de: EE.UU., Europa, China, Latam, Asia, África, Global"""

def file_to_base64(file_bytes):
    return base64.b64encode(file_bytes).decode('utf-8')

def extract_from_pdf(api_key, file_bytes, filename):
    """Extrae datos de un PDF de statement usando Claude."""
    client = anthropic.Anthropic(api_key=api_key)
    b64 = file_to_base64(file_bytes)

    response = client.messages.create(
        model="claude-3-5-sonnet-20241022",
        max_tokens=4000,
        system=SYSTEM_PROMPT,
        messages=[{
            "role": "user",
            "content": [
                {"type": "document",
                 "source": {"type": "base64", "media_type": "application/pdf", "data": b64}},
                {"type": "text",
                 "text": f"Extraé los datos financieros de este statement. Archivo: {filename}. Respondé solo con JSON válido."}
            ]
        }]
    )

    text = response.content[0].text.strip()
    text = re.sub(r'^```json?\n?', '', text)
    text = re.sub(r'\n?```$', '', text)
    return json.loads(text.strip())

def extract_from_excel(file_bytes):
    """Extrae posiciones del Excel Unrealized Gain/Loss de Pershing."""
    import io
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # Find the right sheet
    ws = None
    for sheet_name in wb.sheetnames:
        if 'export' in sheet_name.lower() or 'unrealized' in sheet_name.lower():
            ws = wb[sheet_name]
            break
    if ws is None:
        ws = wb.active

    VALID_TYPES = {
        'Common Stocks', 'American Depository Receipt / Global Depository Receipt (Common)',
        'U.S. Treasury Securities', 'Corporate Bonds', 'Exchange-Traded Funds',
        'Exchange-Traded Notes', 'Mutual Funds', 'Money Market'
    }

    TYPE_MAP = {
        'Common Stocks': ('Acciones', 'Renta Variable'),
        'American Depository Receipt / Global Depository Receipt (Common)': ('ADR', 'Renta Variable'),
        'U.S. Treasury Securities': ('Bono Gobierno', 'Renta Fija'),
        'Corporate Bonds': ('Bono Corporativo', 'Renta Fija'),
        'Exchange-Traded Funds': ('ETF', 'Alternativo'),
        'Exchange-Traded Notes': ('ETN', 'Alternativo'),
        'Mutual Funds': ('Fondo Mutuo', 'Fondos'),
        'Money Market': ('Money Market', 'Efectivo'),
    }

    GEO_MAP = {
        'INTC':'EE.UU.','GOOG':'EE.UU.','MSFT':'EE.UU.','UNH':'EE.UU.',
        'JNJ':'EE.UU.','BRK B':'EE.UU.','WMT':'EE.UU.','AMZN':'EE.UU.',
        'KO':'EE.UU.','CPNG':'Asia','BABA':'China','ASML':'Europa',
        'MELI':'Latam','JMIA':'África','BYMAF':'Latam',
    }
    SECTOR_MAP = {
        'INTC':'Tecnología','GOOG':'Tecnología','MSFT':'Tecnología','ASML':'Tecnología',
        'UNH':'Salud','JNJ':'Salud',
        'BRK B':'Financiero','BYMAF':'Financiero',
        'WMT':'Cons. Básico','KO':'Cons. Básico',
        'MELI':'Cons. Discrecional','BABA':'Cons. Discrecional',
        'AMZN':'Cons. Discrecional','JMIA':'Cons. Discrecional','CPNG':'Cons. Discrecional',
    }

    agg = defaultdict(lambda: {
        'desc':'','stype':'','acat':'','geo':'Global','sector':'',
        'qty':0,'price':0,'mv':0,'cost':0
    })

    for row in ws.iter_rows(min_row=15, values_only=True):
        stype_raw = row[0]
        if stype_raw not in VALID_TYPES:
            continue
        sym   = str(row[49] or row[1] or '').strip()
        desc  = str(row[2] or '').strip()[:40]
        qty   = float(row[10] or 0)
        price = float(row[17] or 0)
        mv    = float(row[7]  or 0)
        cost  = float(row[6]  or 0)

        if mv == 0 and cost == 0:
            continue

        stype, acat = TYPE_MAP.get(stype_raw, ('Acciones','Renta Variable'))
        agg[sym]['desc']   = desc
        agg[sym]['stype']  = stype
        agg[sym]['acat']   = acat
        agg[sym]['geo']    = GEO_MAP.get(sym, 'EE.UU.' if acat == 'Renta Variable' else 'Global')
        agg[sym]['sector'] = SECTOR_MAP.get(sym, acat)
        agg[sym]['qty']   += qty
        agg[sym]['price']  = price
        agg[sym]['mv']    += mv
        agg[sym]['cost']  += cost

    positions = []
    for sym, d in agg.items():
        if d['mv'] == 0:
            continue
        gl    = d['mv'] - d['cost']
        positions.append({
            'symbol':      sym,
            'description': d['desc'],
            'secType':     d['stype'],
            'assetCategory': d['acat'],
            'geo':         d['geo'],
            'sector':      d['sector'],
            'quantity':    d['qty'],
            'price':       d['price'],
            'marketValue': d['mv'],
            'costBasis':   d['cost'],
            'gainLoss':    gl,
            'gainLossPct': gl/d['cost']*100 if d['cost'] else None,
        })

    return sorted(positions, key=lambda x: -x['marketValue'])

def merge_extracted_data(existing_history, new_data):
    """
    Fusiona datos extraídos con el historial existente.
    existing_history: dict {year_key: {vi, dep, ret, vf, res, r}}
    new_data: resultado del extractor (puede ser annual o monthly)
    Retorna: historial actualizado
    """
    history = dict(existing_history)

    if new_data.get('type') == 'annual' and 'years' in new_data:
        for yr, vals in new_data['years'].items():
            history[int(yr)] = vals
    elif new_data.get('type') == 'monthly':
        label = new_data.get('label', f"YTD {new_data.get('year', '')}")
        history[label] = {
            'vi':  new_data.get('vi', 0),
            'dep': new_data.get('dep', 0),
            'ret': new_data.get('ret', 0),
            'vf':  new_data.get('vf', 0),
            'res': new_data.get('res', 0),
            'r':   new_data.get('r', 0.0),
        }
    return history
